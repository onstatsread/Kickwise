"""
Kickwise Backend — FastAPI server
Primary market-odds source: AnnaBet
Fallback market-odds sources are kept available but AnnaBet is tried first.

Deploy to Render.com.
"""

from fastapi import FastAPI, Query
from fastapi.middleware.cors import CORSMiddleware
import requests, os, subprocess, statistics, tempfile, shutil, difflib, re, time, calendar, hashlib
from scipy.stats import poisson
from datetime import date
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor

from annabet_leagues import ANNABET_LEAGUE_IDS

# Existing fallbacks. AnnaBet is now the PRIMARY odds source.
from odds import get_odds_for_card, get_ou25_for_card, router as odds_router
from api_football import get_fallback_odds
from odds_api_io import get_odds_api_io_fallback, get_ou25_api_io_fallback

# NEW — Oddsbook fixtures+odds fetcher, used ONLY by the /predict-oddsbook-test
# endpoint for side-by-side comparison. Does not touch the existing AnnaBet
# fixtures/predict flow at all.
from oddsbook_odds import get_fixtures_for_day, get_oddsbook_market_odds
from datetime import datetime as _datetime


app = FastAPI(title="Kickwise API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["GET"],
    allow_headers=["*"],
)

app.include_router(odds_router)


# ============================================================
# CONSTANTS / SESSIONS
# ============================================================

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/124.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.soccerstats.com/",
    "Connection": "keep-alive",
}

BASE = "https://www.soccerstats.com"
MODEL = "A_mix2.xlsx"

ANNABET_HEADERS = {
    "User-Agent": HEADERS["User-Agent"],
    "Accept": HEADERS["Accept"],
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate",
    "Connection": "keep-alive",
    "Referer": "https://annabet.com/en/soccerstats/",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "same-origin",
    "Upgrade-Insecure-Requests": "1",
}

ANNABET_SESSION = requests.Session()
ANNABET_SESSION.headers.update(ANNABET_HEADERS)

ANNABET_LOGIN_URL = "https://annabet.com/auth/ASEngine/ASAjax.php"
ANNABET_USERNAME = os.environ.get("ANNABET_USERNAME", "")
ANNABET_PASSWORD = os.environ.get("ANNABET_PASSWORD", "")

_ANNABET_LOGGED_IN = False


def annabet_login():
    global _ANNABET_LOGGED_IN

    if not ANNABET_USERNAME or not ANNABET_PASSWORD:
        print("AnnaBet login skipped — ANNABET_USERNAME/ANNABET_PASSWORD not set")
        return False

    password_hash = hashlib.sha512(ANNABET_PASSWORD.encode("utf-8")).hexdigest()

    try:
        resp = ANNABET_SESSION.post(
            ANNABET_LOGIN_URL,
            data={
                "action": "checkLogin",
                "username": ANNABET_USERNAME,
                "password": password_hash,
            },
            timeout=20,
        )
        resp.raise_for_status()
        result = resp.json()

        if result.get("page"):
            _ANNABET_LOGGED_IN = True
            print("AnnaBet login succeeded")
            return True

        print(f"AnnaBet login failed — unexpected response: {result}")
        return False

    except Exception as e:
        print(f"AnnaBet login failed: {e}")
        return False


annabet_login()

ANNABET_TABLE_HEADER = [
    "#", "Team", "GP", "W", "T", "L", "GF", "GA", "Diff",
    "Pts", "Pts/G", "W%", "ØGF", "ØGA"
]

ANNABET_SERIE_ID = {
    "belarus": 232, "brazil": 217, "brazil2": 259, "canada": 750,
    "chile": 301, "china": 248, "china2": 731, "colombia": 329,
    "ecuador": 313, "estonia": 242, "faroeislands": 368, "finland": 7,
    "finland2": 35, "georgia": 235, "iceland": 114, "iceland2": 392,
    "ireland": 42, "ireland2": 163, "kazakhstan": 328, "latvia": 223,
    "lithuania": 226, "malaysia": 521, "norway": 36, "norway2": 173,
    "paraguay": 347, "peru": 321, "southkorea": 249, "southkorea2": 543,
    "sweden": 32, "sweden2": 33, "uruguay": 439, "usa": 43, "usa2": 362,
    "venezuela": 314,
    "englandsouthern": 263, "germany": 8, "belgium": 697,
    "algeria": 257, "australia": 148, "australiabrisbane": 465,
    "chile2": 669, "bolivia": 400, "greece2": 769, "estonia2": 647,
    "iceland3": 688, "greece3": 258, "indiail": 276, "indiaisl": 676,
    "jamaica": 628, "iranazadegan": 513, "kenya": 663, "jordan": 445,
    "morocco": 330, "singapore": 213, "newzealand": 454, "syria": 300,
    "thailand": 356, "vietnam": 426, "taiwan": 763,
    "turkmenistan": 764, "tajikistan": 761,
}

_ANNABET_ID_TO_CODE = {v: k for k, v in ANNABET_SERIE_ID.items()}
ANNABET_UPCOMING_URL = "https://annabet.com/en/soccerstats/upcoming/"

_ANNABET_FIXTURES_CACHE = {}
ANNABET_FIXTURES_CACHE_TTL = 1800

_ANNABET_ODDS_CACHE = {}
ANNABET_ODDS_CACHE_TTL = 300

_ANNABET_SERIE_LINK_RE = re.compile(r"/serie_(\d+)_")
_ANNABET_DATETIME_RE = re.compile(r"(\d{1,2})\.(\d{1,2})\.\s+(\d{1,2}):(\d{2})")

TIME_RE = re.compile(r"\b([01]?\d|2[0-3]):([0-5]\d)\b")
DAY_RE = re.compile(r"^(Mon|Tue|Wed|Thu|Fri|Sat|Sun)\b\s*")


def _annabet_looks_logged_out(html):
    lower = html.lower()
    return (
        "please register" in lower
        or "register or log in" in lower
    ) and "logout" not in lower


def annabet_get(url, **kwargs):
    global _ANNABET_LOGGED_IN

    resp = ANNABET_SESSION.get(url, **kwargs)

    if ANNABET_USERNAME and ANNABET_PASSWORD and _annabet_looks_logged_out(resp.text):
        print("AnnaBet session appears expired — re-authenticating")
        _ANNABET_LOGGED_IN = False

        if annabet_login():
            resp = ANNABET_SESSION.get(url, **kwargs)

    return resp


def fetch_all_upcoming_annabet():
    cached = _ANNABET_FIXTURES_CACHE.get("_all")
    if cached and time.time() - cached[0] < ANNABET_FIXTURES_CACHE_TTL:
        return cached[1]

    resp = annabet_get(ANNABET_UPCOMING_URL, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    by_league = {}

    for row in soup.find_all("tr"):
        cells = row.find_all("td")
        if len(cells) < 3:
            continue

        row_text = cells[0].get_text(" ", strip=True)
        dt_match = _ANNABET_DATETIME_RE.search(row_text)
        if not dt_match:
            continue

        league_link = None
        for cell in cells:
            a = cell.find("a", href=_ANNABET_SERIE_LINK_RE)
            if a:
                league_link = a
                break

        if not league_link:
            continue

        serie_match = _ANNABET_SERIE_LINK_RE.search(league_link["href"])
        if not serie_match:
            continue

        serie_id = int(serie_match.group(1))
        code = _ANNABET_ID_TO_CODE.get(serie_id)
        if not code:
            continue

        team_link = None
        for cell in cells:
            a = cell.find("a", href=re.compile(r"h2h\.php"))
            if a:
                team_link = a
                break

        if not team_link:
            continue

        team_text = team_link.get_text(" ", strip=True)
        if " - " not in team_text:
            continue

        home, away = team_text.split(" - ", 1)

        day, month, hour, minute = dt_match.groups()

        odds_numbers = []
        for cell in cells:
            text = cell.get_text(" ", strip=True)
            for raw in re.findall(r"\b\d+(?:\.\d+)\b", text):
                n = _annabet_float(raw)
                if n is not None:
                    odds_numbers.append(n)

        home_odds = draw_odds = away_odds = None
        if len(odds_numbers) >= 3:
            home_odds, draw_odds, away_odds = odds_numbers[-3:]

        by_league.setdefault(code, []).append({
            "date": f"{day}.{month}.",
            "time": f"{hour}:{minute}",
            "home": home.strip(),
            "away": away.strip(),
            "h2h_url": _absolute_annabet_url(team_link.get("href")),
            "home_odds": home_odds,
            "draw_odds": draw_odds,
            "away_odds": away_odds,
        })

    _ANNABET_FIXTURES_CACHE["_all"] = (time.time(), by_league)
    return by_league


def _absolute_annabet_url(href):
    if not href:
        return None
    if href.startswith("http://") or href.startswith("https://"):
        return href
    if href.startswith("/"):
        return "https://annabet.com" + href
    return "https://annabet.com/en/soccerstats/" + href


def fetch_fixtures_annabet(code, day, month):
    by_league = fetch_all_upcoming_annabet()
    target = f"{day}.{month}."

    return [
        {
            "time": m["time"],
            "home": m["home"],
            "away": m["away"],
            "h2h_url": m.get("h2h_url"),
        }
        for m in by_league.get(code, [])
        if m["date"] == target
    ]


def _annabet_get_header(table):
    rows = table.find_all("tr")
    if not rows:
        return None
    return [c.get_text(strip=True) for c in rows[0].find_all(["td", "th"])]


def _annabet_parse_table(table):
    teams = {}

    for row in table.find_all("tr")[1:]:
        cells = [c.get_text(strip=True) for c in row.find_all("td")]
        if len(cells) < 8:
            continue

        try:
            name = cells[1]
            gp = int(cells[2])
            gf = int(cells[6])
            ga = int(cells[7])

            teams[name] = {
                "gp": gp,
                "gf": gf,
                "ga": ga,
            }
        except (ValueError, IndexError):
            continue

    return teams


def fetch_stats_annabet(serie_id):
    url = f"https://annabet.com/en/soccerstats/serie_{serie_id}_x.html"

    resp = ANNABET_SESSION.get(url, timeout=20)
    resp.raise_for_status()

    soup = BeautifulSoup(resp.text, "html.parser")
    tables = soup.find_all("table")

    matching = [
        t for t in tables
        if _annabet_get_header(t) == ANNABET_TABLE_HEADER
    ]

    if len(matching) < 3:
        return {}

    all_table, home_table, away_table = matching[:3]

    home_data = _annabet_parse_table(home_table)
    away_data = _annabet_parse_table(away_table)

    result = {}

    for team, h in home_data.items():
        a = away_data.get(team)
        if not a:
            continue

        gp = h["gp"] + a["gp"]
        gf = h["gf"] + a["gf"]
        ga = h["ga"] + a["ga"]

        result[team] = {
            "gp": gp,
            "gf": gf / gp if gp else 0,
            "ga": ga / gp if gp else 0,
            "tot": (gf + ga) / gp if gp else 0,
            "hgf": h["gf"] / h["gp"] if h["gp"] else 0,
            "hga": h["ga"] / h["gp"] if h["gp"] else 0,
            "htot": (h["gf"] + h["ga"]) / h["gp"] if h["gp"] else 0,
            "agf": a["gf"] / a["gp"] if a["gp"] else 0,
            "aga": a["ga"] / a["gp"] if a["gp"] else 0,
            "atot": (a["gf"] + a["ga"]) / a["gp"] if a["gp"] else 0,
        }

    return result


_STATS_CACHE = {}
STATS_CACHE_TTL = 3600


def fetch_stats(code):
    cached = _STATS_CACHE.get(code)

    if cached and time.time() - cached[0] < STATS_CACHE_TTL:
        return cached[1]

    if code in ANNABET_SERIE_ID:
        try:
            result = fetch_stats_annabet(ANNABET_SERIE_ID[code])

            if result:
                _STATS_CACHE[code] = (time.time(), result)
                return result

        except Exception as e:
            print(f"AnnaBet stats failed for {code}: {e}")

    return {}


def _annabet_float(value):
    if value is None:
        return None

    s = str(value).strip()

    if not re.fullmatch(r"\d+(?:\.\d+)?", s):
        return None

    try:
        value = float(s)
    except ValueError:
        return None

    if value <= 1.0 or value > 100:
        return None

    return value


def _add_implied_pct(odds_dict, *keys):
    if not odds_dict:
        return odds_dict

    vals = [odds_dict.get(k) for k in keys]

    if any(v is None or v <= 0 for v in vals):
        return odds_dict

    raw = [1 / v for v in vals]
    total = sum(raw)

    if total <= 0:
        return odds_dict

    for k, r in zip(keys, raw):
        pct_key = k.replace("_odds", "_pct")
        odds_dict[pct_key] = round(r / total * 100, 1)

    return odds_dict


def _norm_team(name):
    return " ".join(str(name).lower().split()).strip()


def _team_names_match(a, b):
    a = _norm_team(a)
    b = _norm_team(b)

    return (
        a == b
        or (len(a) >= 5 and a in b)
        or (len(b) >= 5 and b in a)
    )


def _extract_fixture_h2h_and_hda(home, away):
    cached = _ANNABET_ODDS_CACHE.get(("fixture", _norm_team(home), _norm_team(away)))

    if cached and time.time() - cached[0] < ANNABET_ODDS_CACHE_TTL:
        return cached[1]

    try:
        by_league = fetch_all_upcoming_annabet()
    except Exception as e:
        print(f"AnnaBet upcoming odds fetch failed: {e}")
        return None

    target_home = _norm_team(home)
    target_away = _norm_team(away)

    for fixtures in by_league.values():
        for fx in fixtures:
            if _norm_team(fx["home"]) != target_home:
                continue
            if _norm_team(fx["away"]) != target_away:
                continue

            if fx.get("home_odds") is None:
                return None

            result = {
                "home_odds": fx["home_odds"],
                "draw_odds": fx["draw_odds"],
                "away_odds": fx["away_odds"],
                "h2h_url": fx["h2h_url"],
            }

            _ANNABET_ODDS_CACHE[
                ("fixture", target_home, target_away)
            ] = (time.time(), result)

            return result

    return None


def _extract_annabet_ou25(h2h_url):
    if not h2h_url:
        return None

    cache_key = ("ou25", h2h_url)
    cached = _ANNABET_ODDS_CACHE.get(cache_key)

    if cached and time.time() - cached[0] < ANNABET_ODDS_CACHE_TTL:
        return cached[1]

    try:
        resp = annabet_get(
            h2h_url,
            timeout=30
        )
        resp.raise_for_status()
    except Exception as e:
        print(f"AnnaBet H2H O/U fetch failed: {e}")
        return None

    soup = BeautifulSoup(resp.text, "html.parser")

    candidate_tables = []

    for table in soup.find_all("table"):
        text = table.get_text(" ", strip=True).lower()

        if "total goals under-over" not in text:
            continue

        if table.find("a", href=re.compile(r"^/en/link/")):
            candidate_tables.append(table)

    for table in candidate_tables:

        for cell in table.find_all("td", class_="hdr"):

            cell_text = cell.get_text(" ", strip=True)

            if not re.search(
                r"\b2\.5\s+goals\b",
                cell_text,
                re.I
            ):
                continue

            pairs = re.findall(
                r"(\d+(?:\.\d+)?)\s*-\s*(\d+(?:\.\d+)?)",
                cell_text
            )

            for left, right in pairs:
                under = _annabet_float(left)
                over = _annabet_float(right)

                if under is not None and over is not None:
                    result = {
                        "over_odds": over,
                        "under_odds": under,
                    }

                    _ANNABET_ODDS_CACHE[cache_key] = (
                        time.time(),
                        result
                    )

                    return result

    return None


def get_annabet_market_odds(home, away):
    result = {
        "market_odds": None,
        "market_ou25": None,
    }

    fixture = _extract_fixture_h2h_and_hda(home, away)

    if not fixture:
        return result

    result["market_odds"] = _add_implied_pct(
        {
            "home_odds": fixture["home_odds"],
            "draw_odds": fixture["draw_odds"],
            "away_odds": fixture["away_odds"],
        },
        "home_odds", "draw_odds", "away_odds"
    )

    market_ou25 = _extract_annabet_ou25(
        fixture.get("h2h_url")
    )

    result["market_ou25"] = _add_implied_pct(
        market_ou25,
        "over_odds", "under_odds"
    )

    return result


def calc_win_draw_away(lambda_home, lambda_away, max_goals=10):
    try:
        lambda_home = float(lambda_home)
        lambda_away = float(lambda_away)
    except (TypeError, ValueError):
        return None

    if lambda_home < 0 or lambda_away < 0:
        return None

    home_win = draw = away_win = 0.0

    for h in range(max_goals + 1):
        for a in range(max_goals + 1):
            p = poisson.pmf(h, lambda_home) * poisson.pmf(a, lambda_away)

            if h > a:
                home_win += p
            elif h == a:
                draw += p
            else:
                away_win += p

    total = home_win + draw + away_win

    if total <= 0:
        return None

    return {
        "home_pct": round(home_win / total * 100, 1),
        "draw_pct": round(draw / total * 100, 1),
        "away_pct": round(away_win / total * 100, 1),
    }


def pct_to_odds(pct):
    if not pct or pct <= 0:
        return None
    return round(100 / pct, 2)


def calc_odds(lambda_home, lambda_away):
    wda = calc_win_draw_away(lambda_home, lambda_away)

    if not wda:
        return {
            "home_pct": None,
            "draw_pct": None,
            "away_pct": None,
            "home_odds": None,
            "draw_odds": None,
            "away_odds": None,
        }

    return {
        "home_pct": wda["home_pct"],
        "draw_pct": wda["draw_pct"],
        "away_pct": wda["away_pct"],
        "home_odds": pct_to_odds(wda["home_pct"]),
        "draw_odds": pct_to_odds(wda["draw_pct"]),
        "away_odds": pct_to_odds(wda["away_pct"]),
    }


def calc_over_under_25(lambda_home, lambda_away, max_goals=10):
    try:
        lambda_home = float(lambda_home)
        lambda_away = float(lambda_away)
    except (TypeError, ValueError):
        return {
            "over_pct": None,
            "under_pct": None,
            "over_odds": None,
            "under_odds": None,
        }

    if lambda_home < 0 or lambda_away < 0:
        return {
            "over_pct": None,
            "under_pct": None,
            "over_odds": None,
            "under_odds": None,
        }

    over = under = 0.0

    for h in range(max_goals + 1):
        for a in range(max_goals + 1):
            p = poisson.pmf(h, lambda_home) * poisson.pmf(a, lambda_away)

            if h + a > 2.5:
                over += p
            else:
                under += p

    total = over + under

    if total <= 0:
        return {
            "over_pct": None,
            "under_pct": None,
            "over_odds": None,
            "under_odds": None,
        }

    over_pct = round(over / total * 100, 1)
    under_pct = round(under / total * 100, 1)

    return {
        "over_pct": over_pct,
        "under_pct": under_pct,
        "over_odds": pct_to_odds(over_pct),
        "under_odds": pct_to_odds(under_pct),
    }


def resolve_team(name, team_data):
    if name in team_data:
        return name

    substring_candidates = [
        k for k in team_data
        if len(name) >= 5 and (name in k or k in name)
    ]

    if len(substring_candidates) == 1:
        return substring_candidates[0]

    close = difflib.get_close_matches(
        name,
        team_data.keys(),
        n=2,
        cutoff=0.8
    )

    if len(close) == 1:
        return close[0]

    return None


def clean_team_name(name):
    return DAY_RE.sub("", name).strip()


def _norm_key(a, b):
    def clean(s):
        return " ".join(s.lower().split())

    return clean(a), clean(b)


def fetch_fixtures(code, date_str=None):
    if date_str:
        today1 = date_str.strip()
    else:
        d = date.today()
        today1 = f"{d.day} {d.strftime('%b')}"

    if code in ANNABET_SERIE_ID:
        try:
            if not date_str:
                d = date.today()
                return fetch_fixtures_annabet(
                    code,
                    d.day,
                    d.month
                )

            parts = date_str.strip().split()
            day_num = int(parts[0])
            month_num = list(calendar.month_abbr).index(
                parts[1][:3].title()
            )

            return fetch_fixtures_annabet(
                code,
                day_num,
                month_num
            )

        except Exception as e:
            print(f"AnnaBet fixtures failed for {code}: {e}")
            return []

    return []


def run_model(home, away, team_data):

    empty = {
        "d70": "N/A",
        "b120": "N/A",
        "c120": "N/A",
        "b46": "N/A",
        "d64": "N/A",
        "b118": "N/A",
        "aa15": "N/A",
        "b54": "N/A",
        "odds": None,
        "ou25": None,
        "b119": "",
        "d119": "",
        "d70val": "",
        "o73": "",
        "o74": "",
    }

    if home not in team_data or away not in team_data:
        return empty

    data = sorted([
        (
            n,
            d["gp"],
            d["gf"],
            d["ga"],
            d["tot"],
            d["hgf"],
            d["hga"],
            d["htot"],
            d["agf"],
            d["aga"],
            d["atot"]
        )
        for n, d in team_data.items()
    ], key=lambda x: x[0])

    lhs = statistics.mean([d[5] for d in data]) or 1
    lhc = statistics.mean([d[6] for d in data]) or 1
    las = statistics.mean([d[8] for d in data]) or 1
    lac = statistics.mean([d[9] for d in data]) or 1

    wb = load_workbook(MODEL)
    ws = wb.active

    for row in ws.iter_rows(
        min_row=6,
        max_row=42,
        min_col=3,
        max_col=22
    ):
        for cell in row:
            cell.value = None

    for i, d in enumerate(data):
        r = 6 + i

        hs, hc, ht = d[5], d[6], d[7]
        as_, ac, at_ = d[8], d[9], d[10]

        ws.cell(r, 3).value = d[0]
        ws.cell(r, 4).value = d[1]
        ws.cell(r, 5).value = round(d[2], 4)
        ws.cell(r, 6).value = round(d[3], 4)
        ws.cell(r, 7).value = round(d[4], 4)
        ws.cell(r, 8).value = "  "
        ws.cell(r, 9).value = round(hs, 4)
        ws.cell(r, 10).value = round(hc, 4)
        ws.cell(r, 11).value = round(ht, 4)
        ws.cell(r, 12).value = "  "
        ws.cell(r, 13).value = round(as_, 4)
        ws.cell(r, 14).value = round(ac, 4)
        ws.cell(r, 15).value = round(at_, 4)
        ws.cell(r, 16).value = round(hs / lhs, 4)
        ws.cell(r, 17).value = round(hc / lhc, 4)
        ws.cell(r, 18).value = round(as_ / las, 4)
        ws.cell(r, 19).value = round(ac / lac, 4)
        ws.cell(r, 20).value = round(
            max((hs - as_) / d[1], 0),
            4
        )
        ws.cell(r, 22).value = round(
            (ht + at_) / 2,
            4
        )

    ws["B69"] = home
    ws["C69"] = away
    ws.title = "Sheet1"

    tmp_dir = tempfile.mkdtemp()
    tmp_file = os.path.join(tmp_dir, "fm_tmp.xlsx")
    out_dir = os.path.join(tmp_dir, "out")
    os.makedirs(out_dir)

    try:
        wb.save(tmp_file)

        subprocess.run(
            [
                "libreoffice",
                "--headless",
                "--calc",
                "--convert-to",
                "xlsx",
                "--outdir",
                out_dir,
                tmp_file,
            ],
            capture_output=True,
            timeout=90,
        )

        out_file = os.path.join(out_dir, "fm_tmp.xlsx")

        if not os.path.exists(out_file):
            return empty

        wb2 = load_workbook(
            out_file,
            data_only=True
        )

        ws2 = wb2.active

        d70 = str(ws2["D69"].value or "")
        c120 = str(ws2["C120"].value or "")

        b119_raw = str(ws2["B119"].value or "")
        c119_raw = str(ws2["C119"].value or "")
        d119_raw = str(ws2["D119"].value or "")

        parts = [
            x for x in [
                b119_raw,
                c119_raw,
                d119_raw
            ]
            if x and x not in (
                "run",
                "#NAME?",
                "#N/A",
                "None"
            )
        ]

        b120 = " /".join(parts)

        def safe(ref, sheet=None):
            s = sheet if sheet else ws2
            v = str(s[ref].value or "")

            if v in (
                "#NAME?",
                "#N/A",
                "#VALUE!",
                "None"
            ):
                return ""

            return v

        b118_parts = [
            x for x in [
                safe("L115"),
                safe("N111"),
                safe("O111")
            ]
            if x
        ]

        b118 = "/ ".join(b118_parts)

        b46_parts = [
            x for x in [
                safe("C114"),
                safe("O84"),
                safe("O85")
            ]
            if x
        ]

        b46 = ", ".join(b46_parts)

        d64 = safe("D64")

        sheet2 = wb2["Sheet2"]
        aa15 = safe("AA15", sheet2)

        t99 = safe("T99")
        t100 = safe("T100")

        b54_parts = [
            x for x in [t99, t100]
            if x
        ]

        b54 = "/ ".join(b54_parts)

        lambda_home = sheet2["C5"].value
        lambda_away = sheet2["D5"].value

        odds = calc_odds(
            lambda_home,
            lambda_away
        )

        ou25 = calc_over_under_25(
            lambda_home,
            lambda_away
        )

        b119_raw = safe("B119")
        d119_raw = safe("D119")

        b119 = (
            b119_raw
            if b119_raw not in ("run", "")
            else ""
        )

        d119 = (
            d119_raw
            if d119_raw not in ("run", "")
            else ""
        )

        d70_val = safe("D70")
        o73 = safe("O73", sheet2)

        o74_raw = sheet2["O74"].value

        try:
            o74 = (
                str(round(float(o74_raw), 1))
                + "%"
                if o74_raw is not None
                else ""
            )
        except Exception:
            raw = str(o74_raw or "")
            o74 = (
                ""
                if raw in (
                    "#NAME?",
                    "#N/A",
                    "#VALUE!",
                    "None"
                )
                else raw
            )

        return {
            "d70": d70,
            "b120": b120,
            "c120": c120,
            "b46": b46,
            "d64": d64,
            "b118": b118,
            "aa15": aa15,
            "b54": b54,
            "odds": odds,
            "ou25": ou25,
            "b119": b119,
            "d119": d119,
            "d70val": d70_val,
            "o73": o73,
            "o74": o74,
        }

    finally:
        shutil.rmtree(
            tmp_dir,
            ignore_errors=True
        )


@app.get("/predict")
async def predict(
    league: str = Query(...),
    home: str = Query(...),
    away: str = Query(...)
):
    t0 = time.time()

    team_data = fetch_stats(league)

    print(
        f"fetch_stats({league}) took "
        f"{time.time() - t0:.1f}s"
    )

    resolved_h = resolve_team(
        home,
        team_data
    )

    resolved_a = resolve_team(
        away,
        team_data
    )

    h = resolved_h or home
    a = resolved_a or away

    with ThreadPoolExecutor(max_workers=2) as executor:
        f1 = executor.submit(
            run_model,
            h,
            a,
            team_data
        )

        f2 = executor.submit(
            run_model,
            a,
            h,
            team_data
        )

        r1 = f1.result()
        r2 = f2.result()

    print(
        f"run_model({home} - {away}) took "
        f"{time.time() - t0:.1f}s"
    )

    market_odds = None
    market_ou25 = None

    try:
        annabet = get_annabet_market_odds(
            home,
            away
        )

        market_odds = annabet.get(
            "market_odds"
        )

        market_ou25 = annabet.get(
            "market_ou25"
        )

        print(
            f"AnnaBet market odds "
            f"{home} - {away}: "
            f"HDA={market_odds}, "
            f"OU25={market_ou25}"
        )

    except Exception as e:
        print(
            f"AnnaBet market odds failed "
            f"for {home} - {away}: {e}"
        )

    if not market_odds or not market_ou25:
        try:
            sport_key = None

            env_key = (
                "ODDS_SPORT_KEY_"
                + league.upper()
            )

            sport_key = os.environ.get(
                env_key
            )

            if sport_key:

                if not market_odds:
                    market_odds = await get_odds_for_card(
                        sport_key,
                        home,
                        away
                    )

                if not market_ou25:
                    market_ou25 = await get_ou25_for_card(
                        sport_key,
                        home,
                        away
                    )

        except Exception as e:
            print(
                f"The Odds API fallback failed: {e}"
            )

    if not market_odds:
        try:
            market_odds = (
                await get_odds_api_io_fallback(
                    home,
                    away
                )
            )
        except Exception as e:
            print(
                f"Odds-API.io HDA fallback failed: {e}"
            )

    if not market_ou25:
        try:
            market_ou25 = (
                await get_ou25_api_io_fallback(
                    home,
                    away
                )
            )
        except Exception as e:
            print(
                f"Odds-API.io OU25 fallback failed: {e}"
            )

    if not market_odds:
        try:
            market_odds = await get_fallback_odds(
                home,
                away
            )
        except Exception as e:
            print(
                f"API-Football fallback failed: {e}"
            )

    value_pct = None
    value_signal = None

    model_odds = r1.get("odds")

    if model_odds and market_odds:

        def pct_diff(market_o, model_o):
            if not model_o or not market_o:
                return None

            return round(
                ((market_o - model_o) / model_o) * 100,
                1
            )

        home_v = pct_diff(
            market_odds.get("home_odds"),
            model_odds.get("home_odds")
        )

        draw_v = pct_diff(
            market_odds.get("draw_odds"),
            model_odds.get("draw_odds")
        )

        away_v = pct_diff(
            market_odds.get("away_odds"),
            model_odds.get("away_odds")
        )

        value_pct = {
            "home": home_v,
            "draw": draw_v,
            "away": away_v,
        }

        decision = ""
        under_flag = ""

        if (
            home_v is not None
            and away_v is not None
        ):

            total_v = round(
                sum(
                    x for x in [
                        home_v,
                        draw_v,
                        away_v
                    ]
                    if x is not None
                ),
                1
            )

            value_pct["total"] = total_v

            abs_sum = (
                abs(home_v)
                + abs(draw_v or 0)
                + abs(away_v)
            )

            signal = ""

            if abs_sum > 0:

                home_share = (
                    abs(home_v)
                    / abs_sum
                    * 100
                )

                away_share = (
                    abs(away_v)
                    / abs_sum
                    * 100
                )

                share_diff = (
                    home_share
                    - away_share
                )

                value_pct["share_diff"] = round(
                    share_diff,
                    1
                )

                if share_diff > 0:
                    signal = "Away"

                elif share_diff < 0:
                    signal = "Home"

            if (
                signal == "Away"
                and away_v < 0
            ):
                decision = "Away"

            elif (
                signal == "Home"
                and home_v < 0
            ):
                decision = "Home"

            else:

                if home_v < 0 and away_v < 0:
                    decision = (
                        "Home 2-handicap"
                        if home_v < away_v
                        else "Away 2-handicap"
                    )

                elif home_v < 0:
                    decision = "Home 2-handicap"

                elif away_v < 0:
                    decision = "Away 2-handicap"

            same_sign = (
                (home_v < 0 and away_v < 0)
                or
                (home_v > 0 and away_v > 0)
            )

            under_flag = (
                ""
                if same_sign
                else (
                    "under"
                    if -30 <= total_v <= 0
                    else ""
                )
            )

        value_signal = {
            "decision": decision,
            "under": under_flag
        }

    ou25_value_pct = None
    ou25_value_signal = None
    prediction_3 = ""
    prediction_3_gate = False

    model_ou25 = r1.get("ou25")

    if model_ou25 and market_ou25:

        def pct_diff_ou(market_o, model_o):
            if not model_o or not market_o:
                return None

            return round(
                ((market_o - model_o) / model_o) * 100,
                1
            )

        over_v = pct_diff_ou(
            market_ou25.get("over_odds"),
            model_ou25.get("over_odds")
        )

        under_v = pct_diff_ou(
            market_ou25.get("under_odds"),
            model_ou25.get("under_odds")
        )

        if (
            over_v is not None
            or under_v is not None
        ):

            ou_total_v = round(
                sum(
                    x for x in [
                        over_v,
                        under_v
                    ]
                    if x is not None
                ),
                1
            )

            ou_abs_sum = (
                abs(over_v or 0)
                + abs(under_v or 0)
            )

            ou_abs_diff = (
                abs(over_v or 0)
                - abs(under_v or 0)
            )

            def ou_share(v):
                if (
                    v is None
                    or ou_abs_sum == 0
                ):
                    return None

                return round(
                    abs(v)
                    / ou_abs_sum
                    * 100,
                    1
                )

            over_share_v = ou_share(
                over_v
            )

            under_share_v = ou_share(
                under_v
            )

            share_diff_v = round(
                (over_share_v or 0)
                - (under_share_v or 0),
                1
            )

            ou25_value_pct = {
                "over": over_v,
                "under": under_v,
                "total": ou_total_v,
                "over_share": over_share_v,
                "under_share": under_share_v,
                "abs_diff": round(
                    ou_abs_diff,
                    1
                ),
                "share_diff": share_diff_v,
            }

            if ou_abs_diff > 0:
                step4 = "over"
            elif ou_abs_diff < 0:
                step4 = "under"
            else:
                step4 = ""

            step5 = ""

            cv = over_v or 0
            dv = under_v or 0

            if (
                step4 == "under"
                and dv < 0
                and cv > 0
            ):
                step5 = "under"

            elif (
                step4 == "over"
                and cv < 0
                and dv > 0
            ):
                step5 = "over"

            elif cv < 0 and dv < 0:
                step5 = (
                    "over+"
                    if cv < dv
                    else "under+"
                )

            elif cv < 0:
                step5 = "over+"

            elif dv < 0:
                step5 = "under+"

            same_sign = (
                (cv < 0 and dv < 0)
                or
                (cv > 0 and dv > 0)
            )

            step6 = (
                ""
                if same_sign
                else (
                    "under"
                    if -30 <= ou_total_v <= 0
                    else ""
                )
            )

            if (
                step4 == "under"
                and step6 == "under"
            ):
                ou_result_signal = "under confirmed"

            elif (
                step4 != "under"
                and step6 == "under"
            ):
                ou_result_signal = "under"

            else:
                ou_result_signal = "over"

            ou25_value_signal = {
                "result": ou_result_signal,
                "step4": step4,
                "step5": step5,
                "step6": step6,
            }

            if step4:

                hda_decision = (
                    value_signal or {}
                ).get(
                    "decision",
                    ""
                ).lower()

                if step5 in (
                    "over+",
                    "over"
                ):
                    prediction_3 = (
                        "Home"
                        if "home" in hda_decision
                        else "Home handicap"
                    )

                elif step5 in (
                    "under+",
                    "under"
                ):
                    prediction_3 = (
                        "Away"
                        if "away" in hda_decision
                        else "Away handicap"
                    )

            hda_total = (
                value_pct or {}
            ).get("total")

            hda_under = (
                value_signal or {}
            ).get(
                "under",
                ""
            )

            if (
                ou_total_v is not None
                and -20 <= ou_total_v <= 0
                and ou_result_signal in (
                    "under",
                    "under confirmed"
                )
                and hda_total is not None
                and -20 <= hda_total <= 0
                and hda_under == "under"
            ):

                prediction_3_gate = True

                decision_base = (
                    value_signal or {}
                ).get(
                    "decision",
                    ""
                )

                b46_combined = (
                    (r1.get("b46") or "")
                    + " "
                    + (r2.get("b46") or "")
                ).lower()

                b46_match = re.search(
                    r"(\d+\s*goals)",
                    b46_combined
                )

                b46_goals = (
                    b46_match.group(1)
                    .replace(" ", "")
                    if b46_match
                    else ""
                )

                suffix = (
                    f"under {b46_goals}"
                    if b46_goals
                    else "under"
                )

                prediction_3 = (
                    f"{decision_base}/ {suffix}"
                    if decision_base
                    else suffix.capitalize()
                )

    return {
        "home": h,
        "away": a,

        "d70": r1["d70"],
        "b120": r1["b120"],
        "c120": r1["c120"],
        "b46": r1["b46"],
        "d64": r1["d64"],
        "b118": r1["b118"],
        "aa15": r1["aa15"],
        "b54": r1["b54"],

        "odds": r1.get("odds"),
        "ou25": r1.get("ou25"),

        "market_ou25": market_ou25,
        "market_odds": market_odds,

        "ou25_value_pct": ou25_value_pct,
        "ou25_value_signal": ou25_value_signal,

        "prediction_3": prediction_3,
        "prediction_3_gate": prediction_3_gate,

        "value_pct": value_pct,
        "value_signal": value_signal,

        "b119": r1["b119"],
        "d119": r1["d119"],
        "d70val": r1["d70val"],
        "o73": r1["o73"],
        "o74": r1["o74"],

        "d70r": r2["d70"],
        "b120r": r2["b120"],
        "c120r": r2["c120"],
        "b46r": r2["b46"],
        "d64r": r2["d64"],
        "b118r": r2["b118"],
        "aa15r": r2["aa15"],
        "b54r": r2["b54"],

        "oddsr": r2.get("odds"),

        "b119r": r2["b119"],
        "d119r": r2["d119"],
        "d70valr": r2["d70val"],
        "o73r": r2["o73"],
        "o74r": r2["o74"],
    }


# ============================================================
# NEW — TEST ENDPOINT: Oddsbook fixtures+odds, AnnaBet stats/model
#
# Uses the SAME fetch_stats()/run_model() as the live /predict above
# (untouched) but sources fixtures + market odds from Oddsbook instead
# of AnnaBet, for side-by-side comparison. league_code is an AnnaBet
# short code (e.g. "brazil") since that's still what fetch_stats()
# needs — team names get resolved the same way either source names them.
# ============================================================

@app.get("/predict-oddsbook-test")
async def predict_oddsbook_test(
    league: str = Query(..., description="AnnaBet short code, e.g. 'brazil' — used for fetch_stats() only"),
    home: str = Query(...),
    away: str = Query(...),
):
    t0 = time.time()

    team_data = fetch_stats(league)

    resolved_h = resolve_team(home, team_data)
    resolved_a = resolve_team(away, team_data)

    h = resolved_h or home
    a = resolved_a or away

    with ThreadPoolExecutor(max_workers=2) as executor:
        f1 = executor.submit(run_model, h, a, team_data)
        f2 = executor.submit(run_model, a, h, team_data)
        r1 = f1.result()
        r2 = f2.result()

    print(f"[oddsbook-test] run_model({home} - {away}) took {time.time() - t0:.1f}s")

    # Oddsbook market odds — replaces get_annabet_market_odds() entirely
    # for this test endpoint. Uses today's date by default; pass
    # ?date=YYYY-MM-DD if testing a future fixture.
    market_odds = None
    market_ou25 = None  # Oddsbook O/U 2.5 still unconfirmed — see oddsbook_odds.py

    try:
        oddsbook_result = get_oddsbook_market_odds(home, away)
        market_odds = oddsbook_result.get("market_odds")
        market_ou25 = oddsbook_result.get("market_ou25")
        print(f"[oddsbook-test] Oddsbook market odds {home} - {away}: HDA={market_odds}")
    except Exception as e:
        print(f"[oddsbook-test] Oddsbook market odds failed for {home} - {away}: {e}")

    model_odds = r1.get("odds")
    value_pct = None
    value_signal = None

    if model_odds and market_odds:
        def pct_diff(market_o, model_o):
            if not model_o or not market_o:
                return None
            return round(((market_o - model_o) / model_o) * 100, 1)

        home_v = pct_diff(market_odds.get("home_odds"), model_odds.get("home_odds"))
        draw_v = pct_diff(market_odds.get("draw_odds"), model_odds.get("draw_odds"))
        away_v = pct_diff(market_odds.get("away_odds"), model_odds.get("away_odds"))

        value_pct = {"home": home_v, "draw": draw_v, "away": away_v}

        decision = ""
        if home_v is not None and away_v is not None:
            if home_v < 0 and away_v < 0:
                decision = "Home 2-handicap" if home_v < away_v else "Away 2-handicap"
            elif home_v < 0:
                decision = "Home 2-handicap"
            elif away_v < 0:
                decision = "Away 2-handicap"

        value_signal = {"decision": decision}

    return {
        "source": "oddsbook (fixtures/odds) + annabet (stats/model)",
        "home": h,
        "away": a,
        "d70": r1["d70"],
        "b120": r1["b120"],
        "c120": r1["c120"],
        "odds": r1.get("odds"),
        "ou25": r1.get("ou25"),
        "market_odds": market_odds,
        "market_ou25": market_ou25,
        "value_pct": value_pct,
        "value_signal": value_signal,
        "prediction_3": "",
    }


@app.get("/fixtures")
def fixtures_endpoint(
    league: str = Query(...),
    date: str = Query(None)
):
    t0 = time.time()

    matches = fetch_fixtures(
        league,
        date
    )

    print(
        f"fetch_fixtures({league}) "
        f"took {time.time() - t0:.1f}s"
    )

    return {
        "league": league,
        "matches": matches
    }


# ============================================================
# NEW — TEST ENDPOINT: raw Oddsbook fixtures for a day, any league,
# no AnnaBet involved. Useful to browse what's available before
# picking a match to run through /predict-oddsbook-test.
# ============================================================

@app.get("/fixtures-oddsbook-test")
def fixtures_oddsbook_test(date_str: str = Query(None, alias="date")):
    target = (
        _datetime.strptime(date_str, "%Y-%m-%d").date()
        if date_str else date.today()
    )
    by_league = get_fixtures_for_day(target)
    return {"date": str(target), "league_count": len(by_league), "leagues": by_league}


@app.get("/health")
def health():
    return {
        "status": "ok"
    }
