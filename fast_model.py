"""
fast_model.py — validated Python replacement for the LibreOffice-based
run_model() in main.py. Uses the `formulas` library to execute the ACTUAL
A_mix2.xlsx formulas directly (not a hand-reimplementation), so results
are mathematically identical to the Excel/LibreOffice pipeline — this was
validated field-by-field against true LibreOffice output on two different
real match datasets before being wired in here.

Why this is safe: every text/number field below was cross-checked against
`libreoffice --headless --convert-to xlsx` on the same inputs and matched
exactly (including Excel error sentinels like "#NAME?" and "run", which
the existing safe()/filtering logic in main.py already knows how to
handle — nothing downstream needs to change).

Usage:
    from fast_model import run_model_fast
    result = run_model_fast(home, away, team_data)
    # same return shape as the old run_model()

First call after import pays a ~15-25s one-time cost to parse the base
workbook. Every call after that (recalculating with new team stats) takes
roughly 5-7s — no subprocess, no LibreOffice, no temp files.
"""
import statistics
import threading
import formulas
from openpyxl.utils import get_column_letter

MODEL_PATH = "A_mix2.xlsx"
_SHEET1 = "'[A_mix2.xlsx]SHEET1'!"

# The base parsed model is expensive (~15-25s) but only needs to happen
# once per process — cache it globally and guard with a lock so concurrent
# requests on server startup don't each try to parse it simultaneously.
_xl_model = None
_model_lock = threading.Lock()


def _get_model():
    global _xl_model
    if _xl_model is None:
        with _model_lock:
            if _xl_model is None:  # re-check inside the lock
                _xl_model = formulas.ExcelModel().loads(MODEL_PATH).finish()
    return _xl_model


def _build_inputs(team_data, home, away):
    """Builds the full input-cell override dict for one prediction call.
    Mirrors run_model()'s exact write pattern: clears the whole 37-row
    grid first (empty string, NOT None — None causes RANK()/VLOOKUP() to
    misbehave on the cleared rows, confirmed during validation), then
    writes real team rows on top, including the same pre-computed/rounded
    ratio columns (P-V) the old code wrote directly rather than letting
    Excel formulas compute them fresh.
    """
    data = sorted([
        (n, d["gp"], d["gf"], d["ga"], d["tot"],
         d["hgf"], d["hga"], d["htot"], d["agf"], d["aga"], d["atot"])
        for n, d in team_data.items()], key=lambda x: x[0])

    lhs = statistics.mean([d[5] for d in data]) or 1
    lhc = statistics.mean([d[6] for d in data]) or 1
    las = statistics.mean([d[8] for d in data]) or 1
    lac = statistics.mean([d[9] for d in data]) or 1

    inputs = {}
    # Clear the full grid first — critical, do not skip. Leftover template
    # data or a prior request's teams sitting in unrelated rows will
    # silently pollute league-average formulas (I44, J44, etc.) that scan
    # the whole C6:C42-style ranges.
    for r in range(6, 43):
        for col in range(3, 23):
            inputs[f"{_SHEET1}{get_column_letter(col)}{r}"] = ""

    for i, d in enumerate(data):
        r = 6 + i
        hs, hc, ht = d[5], d[6], d[7]
        as_, ac, at_ = d[8], d[9], d[10]
        vals = {
            3: d[0], 4: d[1], 5: round(d[2], 4), 6: round(d[3], 4), 7: round(d[4], 4),
            8: "  ",
            9: round(hs, 4), 10: round(hc, 4), 11: round(ht, 4),
            12: "  ",
            13: round(as_, 4), 14: round(ac, 4), 15: round(at_, 4),
            16: round(hs / lhs, 4), 17: round(hc / lhc, 4),
            18: round(as_ / las, 4), 19: round(ac / lac, 4),
            20: round(max((hs - as_) / d[1], 0), 4),
            22: round((ht + at_) / 2, 4),
        }
        for col, val in vals.items():
            inputs[f"{_SHEET1}{get_column_letter(col)}{r}"] = val

    inputs[f"{_SHEET1}B69"] = home
    inputs[f"{_SHEET1}C69"] = away
    return inputs


def _get(solution, sheet, cell):
    """Extract one cell's value from a formulas-library solution object,
    normalizing to a plain Python str/number the same way the old
    openpyxl-based reading code expected (str(cell.value or ""))."""
    target = f"{sheet.upper()}'!{cell.upper()}"
    for k in solution.keys():
        if k.upper().endswith(target):
            v = solution[k].value
            try:
                v = v[0, 0] if hasattr(v, "shape") else v
            except Exception:
                pass
            return v
    return None


def _safe_str(v):
    """Same semantics as main.py's existing safe() helper: converts a raw
    cell value to a clean string, treating Excel error sentinels as
    empty. Handles both plain-string errors and formulas-library's
    XlError objects (both stringify to the same "#NAME?" etc. text)."""
    s = str(v) if v is not None else ""
    return "" if s in ("#NAME?", "#N/A", "#VALUE!", "None") else s


def run_model_fast(home, away, team_data):
    """Drop-in replacement for main.py's run_model(). Same signature,
    same return shape. Raises the same "all N/A" pattern (via the
    caller's existing check) when a team isn't in team_data — this
    function assumes home/away have ALREADY been validated to exist in
    team_data, same precondition the old run_model() had."""
    if home not in team_data or away not in team_data:
        return {"d70": "N/A", "b120": "N/A", "c120": "N/A", "b46": "N/A", "d64": "N/A",
                "b118": "N/A", "aa15": "N/A", "b54": "N/A", "odds": None, "ou25": None,
                "b119": "", "d119": "", "d70val": "", "o73": "", "o74": ""}

    model = _get_model()
    inputs = _build_inputs(team_data, home, away)
    solution = model.calculate(inputs=inputs)

    d70 = _safe_str(_get(solution, "SHEET1", "D69"))
    c120 = _safe_str(_get(solution, "SHEET1", "C120"))

    b119_raw = _safe_str(_get(solution, "SHEET1", "B119"))
    c119_raw = _safe_str(_get(solution, "SHEET1", "C119"))
    d119_raw = _safe_str(_get(solution, "SHEET1", "D119"))
    parts = [x for x in [b119_raw, c119_raw, d119_raw] if x and x not in ("run", "#NAME?", "#N/A", "None")]
    b120 = " /".join(parts)

    l115 = _safe_str(_get(solution, "SHEET1", "L115"))
    n111 = _safe_str(_get(solution, "SHEET1", "N111"))
    o111 = _safe_str(_get(solution, "SHEET1", "O111"))
    b118_parts = [x for x in [l115, n111, o111] if x]
    b118 = "/ ".join(b118_parts)

    c114 = _safe_str(_get(solution, "SHEET1", "C114"))
    o84 = _safe_str(_get(solution, "SHEET1", "O84"))
    o85 = _safe_str(_get(solution, "SHEET1", "O85"))
    b46_parts = [x for x in [c114, o84, o85] if x]
    b46 = ", ".join(b46_parts)

    d64 = _safe_str(_get(solution, "SHEET1", "D64"))

    aa15 = _safe_str(_get(solution, "SHEET2", "AA15"))

    t99 = _safe_str(_get(solution, "SHEET1", "T99"))
    t100 = _safe_str(_get(solution, "SHEET1", "T100"))
    b54_parts = [x for x in [t99, t100] if x]
    b54 = "/ ".join(b54_parts)

    lambda_home = _get(solution, "SHEET2", "C5")
    lambda_away = _get(solution, "SHEET2", "D5")

    # Reuse the SAME Poisson odds/ou25 math main.py already has — these
    # only ever depended on lambda_home/lambda_away, never on the Excel
    # decision tree, so nothing changes here.
    from main import calc_odds, calc_over_under_25
    odds = calc_odds(lambda_home, lambda_away)
    ou25 = calc_over_under_25(lambda_home, lambda_away)

    b119 = b119_raw if b119_raw not in ("run", "") else ""
    d119 = d119_raw if d119_raw not in ("run", "") else ""
    d70_val = _safe_str(_get(solution, "SHEET1", "D70"))

    o73 = _safe_str(_get(solution, "SHEET2", "O73"))
    o74_raw = _get(solution, "SHEET2", "O74")
    try:
        o74 = str(round(float(o74_raw), 1)) + "%" if o74_raw is not None else ""
    except (TypeError, ValueError):
        o74 = _safe_str(o74_raw)

    return {"d70": d70, "b120": b120, "c120": c120, "b46": b46, "d64": d64,
            "b118": b118, "aa15": aa15, "b54": b54, "odds": odds, "ou25": ou25,
            "b119": b119, "d119": d119, "d70val": d70_val,
            "o73": o73, "o74": o74}


def warm_up():
    """Call this once at server startup (e.g. FastAPI's @app.on_event
    'startup' hook) to pay the ~15-25s base-parse cost during deploy/boot
    rather than on a user's first request."""
    _get_model()
